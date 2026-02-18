import json
import logging
import math
import sys
import typing
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class Tooltip:
    def __init__(self, widget: tk.Widget, text_getter: typing.Callable[[], str]) -> None:
        self.widget = widget
        self.text_getter = text_getter
        self.tipwindow: tk.Toplevel | None = None
        self.widget.bind("<Enter>", self._show_tip)
        self.widget.bind("<Leave>", self._hide_tip)

    def _show_tip(self, event: tk.Event | None) -> None:
        text = self.text_getter()
        if not text:
            return
        if self.tipwindow is not None:
            self._hide_tip(None)
        x = y = 0
        if event is not None:
            x = event.x_root + 12
            y = event.y_root + 8
        else:
            x = self.widget.winfo_rootx() + 12
            y = self.widget.winfo_rooty() + 8
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(
            tw,
            text=text,
            justify=tk.LEFT,
            background="#ffffe0",
            relief=tk.SOLID,
            borderwidth=1,
            font=("TkDefaultFont", 9),
            wraplength=280,
        )
        label.pack(ipadx=4, ipady=2)

    def _hide_tip(self, _: tk.Event | None) -> None:
        if self.tipwindow is not None:
            self.tipwindow.destroy()
            self.tipwindow = None


logging.basicConfig(
    filename="cable_calc.log",
    level=logging.ERROR,
    format="%(asctime)s %(levelname)s %(message)s",
)


class CableCalcApp(tk.Tk):
    WINDOW_TITLE_KEY = "app.title"
    WINDOW_GEOMETRY = "1200x800"
    DEFAULT_LANGUAGE = "ru"

    # String constants to avoid duplication
    CIRCUIT_KEY = "Strujni krug"
    PI_KEY = "Pi"
    DELTA_U_KEY = "ΔU %"
    BY_CURRENT_KEY = "По току"
    BY_DELTA_U_KEY = "По ΔU"
    PROTECTION_KEY = "Защита"
    ICALC_KEY = "Icalc [A]"
    TOTAL_DELTA_U_KEY = "Ukupni ΔU %"

    LANGUAGES = {"ru": "Русский", "sr": "Srpski", "en": "English"}
    DEFAULT_MEDIUM = "air"
    DEFAULT_INSTALLATION_METHOD = "C"

    LABEL_KEY_MAP = {
        CIRCUIT_KEY: "label.circuit",
        "Deonica OD": "label.segment_from",
        "Deonica DO": "label.segment_to",
        "Tip-IZOLACIJE": "label.insulation",
        "Tip-PROVODNIKA": "label.conductor",
        "Oznaka-tip-KABLA": "label.cable",
        "Pi, W": "label.pi",
        "Kj": "label.kj",
        "η": "label.eta",
        "Pj": "label.pj",
        "U": "label.voltage",
        "cos φ": "label.cos_phi",
        "Dužina L, m": "label.length",
        "Presek, mm²": "label.area",
        "Način polaganja": "label.installation",
        "Нагруженные жилы (nž)": "label.loaded_cores",
        "Кабелей в группе (для S)": "label.circuits",
        "Параллельные кабели (n∥)": "label.parallel",
        "Среда для Т": "label.medium",
        "Температура, °C": "label.temperature",
        "S": "label.s",
        "T": "label.t",
        "In, A": "label.in",
        "k": "label.k",
        "Ключ ΔU": "label.drop_key",
    }

    RESULT_LABEL_KEY_MAP = {
        "Pj, W": "label.result.pj",
        "Icalc [A]": "label.result.icalc",
        "R_base [Ω/km]": "label.result.rbase",
        "Iz [A]": "label.result.iz",
        "S": "label.result.s",
        "T": "label.result.t",
        DELTA_U_KEY: "label.result.delta",
        TOTAL_DELTA_U_KEY: "label.result.total_delta",
        "Limit ΔU %": "label.result.limit_delta",
        BY_CURRENT_KEY: "label.result.ampacity",
        BY_DELTA_U_KEY: "label.result.drop",
        "Диапазон In [A]": "label.result.in_range",
        "I2 [A]": "label.result.i2",
        PROTECTION_KEY: "label.result.protection",
        "Совместимость IEC": "label.result.compatibility",
        "Рекомендации": "label.result.recommendations",
    }

    TREE_COLUMN_KEYS = {
        CIRCUIT_KEY: "column.circuit",
        "OD": "column.from",
        "DO": "column.to",
        "E": "column.insulation",
        "F": "column.conductor",
        "G": "column.cable",
        "nž": "column.cores",
        "n∥": "column.n_parallel",
        "Кабелей в группе (S)": "column.group_for_s",
        PI_KEY: "column.pi",
        "Kj": "column.kj",
        "η": "column.eta",
        "Pj": "column.pj",
        "U": "column.voltage",
        "cosφ": "column.cos",
        "L": "column.length",
        "Presek": "column.area",
        "Način polaganja": "column.installation",
        "S": "column.s",
        "T": "column.t",
        "In [A]": "column.in",
        "k": "column.k",
        "I2 [A]": "column.i2",
        ICALC_KEY: "column.icalc",
        "R_base [Ω/km]": "column.rbase",
        "ϭ": "column.sigma",
        "Iz [A]": "column.iz",
        DELTA_U_KEY: "column.drop",
        TOTAL_DELTA_U_KEY: "column.total_drop",
        "Limit ΔU %": "column.limit_drop",
        BY_CURRENT_KEY: "column.ampacity",
        BY_DELTA_U_KEY: "column.drop_status",
        PROTECTION_KEY: "column.protection",
        "Ключ": "column.key",
        "Совместимость IEC": "column.compatibility",
    }

    COLUMN_DESC_KEYS = {
        "column.circuit": "label.circuit",
        "column.from": "label.segment_from",
        "column.to": "label.segment_to",
        "column.insulation": "label.insulation",
        "column.conductor": "label.conductor",
        "column.cable": "label.cable",
        "column.cores": "label.loaded_cores",
        "column.n_parallel": "label.parallel",
        "column.group_for_s": "label.circuits",
        "column.pi": "label.pi",
        "column.kj": "label.kj",
        "column.eta": "label.eta",
        "column.pj": "label.pj",
        "column.voltage": "label.voltage",
        "column.cos": "label.cos_phi",
        "column.length": "label.length",
        "column.area": "label.area",
        "column.installation": "label.installation",
        "column.s": "label.s",
        "column.t": "label.t",
        "column.in": "label.in",
        "column.k": "label.k",
        "column.key": "label.drop_key",
        "column.i2": "export.desc.i2",
        "column.icalc": "export.desc.icalc",
        "column.rbase": "export.desc.rbase",
        "column.sigma": "export.desc.sigma",
        "column.iz": "export.desc.iz",
        "column.drop": "export.desc.drop",
        "column.total_drop": "export.desc.total_drop",
        "column.limit_drop": "export.desc.limit_drop",
        "column.ampacity": "export.desc.ampacity",
        "column.drop_status": "export.desc.drop_status",
        "column.protection": "export.desc.protection",
        "column.compatibility": "export.desc.compatibility",
    }

    # Neutralize embedded dictionaries: use external data only
    TRANSLATIONS: dict[str, dict[str, str]] = {}
    TOOLTIPS: dict[str, dict[str, str]] = {}
    HELP_TEXTS: dict[str, str] = {}
    INSULATION_OPTIONS: list[str] = []
    INSULATION_META: dict[str, dict[str, typing.Any]] = {}
    CONDUCTOR_TYPES: list[str] = []
    VOLTAGE_LEVELS: list[str] = []
    TEMPERATURE_MEDIA: dict[str, dict[str, str]] = {}
    INSTALLATION_METHODS: list[str] = []
    STANDARD_SECTIONS: list[float] = []
    METHOD_PREFERENCE: list[str] = []
    STANDARD_CROSS_SECTIONS: list[str] = []
    STANDARD_BREAKER_RATINGS: list[str] = []
    DROP_LIMIT_KEYS: dict[str, float] = {}
    RESISTIVITY_20: dict[str, float] = {}
    TEMP_COEFF: dict[str, float] = {}
    AMPACITY_BASE: dict[str, dict[float, float]] = {}
    AMPACITY_INSULATION_FACTORS: dict[str, dict[str, dict[str, float]]] = {}
    AMPACITY_LOADED_FACTORS: dict[str, dict[int, float]] = {}
    GROUPING_FACTORS: dict[int, float] = {}
    KT_V_TABLE: dict[str, dict[int, float]] = {}
    KT_Z_TABLE: dict[str, dict[int, float]] = {}
    REACTANCE_DATA: dict[str, typing.Any] = {}

    TREE_COLUMNS = (
        CIRCUIT_KEY,
        "OD",
        "DO",
        "E",
        "F",
        "G",
        "nž",
        "n∥",
        "Кабелей в группе (S)",
        PI_KEY,
        "Kj",
        "η",
        "Pj",
        "U",
        "cosφ",
        "L",
        "Presek",
        "Način polaganja",
        "S",
        "T",
        "In [A]",
        "k",
        "I2 [A]",
        ICALC_KEY,
        "R_base [Ω/km]",
        "ϭ",
        "Iz [A]",
        DELTA_U_KEY,
        TOTAL_DELTA_U_KEY,
        "Limit ΔU %",
        BY_CURRENT_KEY,
        BY_DELTA_U_KEY,
        PROTECTION_KEY,
        "Ключ",
        "Совместимость IEC",
    )

    def __init__(self) -> None:
        super().__init__()
        self.geometry(self.WINDOW_GEOMETRY)

        self._language = tk.StringVar(value=self.DEFAULT_LANGUAGE)
        self._language.trace_add("write", self._on_language_change)

        self._text_bindings: list[tuple[typing.Callable[[str], None], str]] = []
        self._menu_text_bindings: list[tuple[tk.Menu, int, str]] = []
        self._tree_heading_bindings: list[tuple[str, str]] = []

        self.style = ttk.Style(self)
        try:
            self.style.theme_use("clam")
        except tk.TclError:
            pass

        default_fg = self.style.lookup("TLabel", "foreground") or "#202020"
        default_bg = self.style.lookup("TLabel", "background") or self.cget("background")
        self.style.configure("ResultKey.TLabel", foreground=default_fg, background=default_bg)
        self.style.configure("ResultValue.TLabel", foreground=default_fg, background=default_bg)
        self.style.configure(
            "ResultAlert.TLabel",
            foreground="#b00020",
            background="#ffe6e6",
        )
        self.style.map("ResultAlert.TLabel", background=[("!disabled", "#ffe6e6")])
        self.style.configure("Invalid.TEntry", fieldbackground="#ffe6e6")
        self.style.map("Invalid.TEntry", fieldbackground=[("!disabled", "#ffe6e6")])
        self.style.configure("Invalid.TCombobox", fieldbackground="#ffe6e6")
        self.style.map(
            "Invalid.TCombobox",
            fieldbackground=[("readonly", "#ffe6e6"), ("!disabled", "#ffe6e6")],
        )

        self._form_values: dict[str, tk.Variable] = {}
        self._input_widgets: dict[str, ttk.Widget] = {}
        self._input_styles: dict[str, str] = {}
        self._combobox_values: dict[str, list[str]] = {}
        self._intermediate_vars: dict[str, tk.StringVar] = {}
        self._intermediate_labels: dict[str, ttk.Label] = {}
        self._table_data: list[dict[str, str]] = []
        self._last_temperature_warning: tuple[str, str, float] | None = None
        self._temperature_editing = False
        self._tooltips: list["Tooltip"] = []
        if self.TEMPERATURE_MEDIA:
            self._medium_selected_key = (
                self.DEFAULT_MEDIUM
                if self.DEFAULT_MEDIUM in self.TEMPERATURE_MEDIA
                else next(iter(self.TEMPERATURE_MEDIA))
            )
        else:
            self._medium_selected_key = self.DEFAULT_MEDIUM
        self._medium_combobox: ttk.Combobox | None = None
        self._help_text_widget: tk.Text | None = None
        self._notebook: ttk.Notebook | None = None
        self._tabs: dict[str, ttk.Frame] = {}
        self._last_icalc: float | None = None
        self._last_result: dict[str, typing.Any] | None = None
        self._voltage_phase_warning_shown = False
        self._loading_project = False
        self._loading_row_into_form = False

        # Load external resources (translations, tooltips, numeric tables)
        self._load_external_resources()

        self._build_menu()
        self._build_layout()

        self.title(self._(self.WINDOW_TITLE_KEY))
        self._apply_language()

    def _(self, key: str) -> str:
        translations = self.TRANSLATIONS.get(key)
        if not translations:
            return key
        language = self._language.get()
        if language in translations:
            return translations[language]
        default_value = translations.get(self.DEFAULT_LANGUAGE)
        if default_value is not None:
            return default_value
        return next(iter(translations.values()))

    def _m(self, key: str, *args: object) -> str:
        """Return translated string with format placeholders {0}, {1}, ... filled."""
        return self._(key).format(*args)

    def _bind_text(self, setter: typing.Callable[[str], None], key: str) -> None:
        self._text_bindings.append((setter, key))
        setter(self._(key))

    def _register_menu_text(self, menu: tk.Menu, index: int, key: str) -> None:
        self._menu_text_bindings.append((menu, index, key))
        menu.entryconfigure(index, label=self._(key))

    def _register_tree_heading(self, column_id: str, key: str) -> None:
        self._tree_heading_bindings.append((column_id, key))
        self.tree.heading(column_id, text=self._(key))

    def _register_notebook_tab(self, tab: ttk.Frame, key: str) -> None:
        if not self._notebook:
            return

        def setter(value: str, tab_ref: ttk.Frame = tab) -> None:
            if self._notebook:
                self._notebook.tab(tab_ref, text=value)

        self._bind_text(setter, key)

    def _on_language_change(self, *_: object) -> None:
        self._apply_language()

    def _apply_language(self) -> None:
        self.title(self._(self.WINDOW_TITLE_KEY))

        for setter, key in self._text_bindings:
            setter(self._(key))

        for menu, index, key in self._menu_text_bindings:
            try:
                menu.entryconfigure(index, label=self._(key))
            except tk.TclError:
                continue

        if hasattr(self, "tree"):
            for column_id, key in self._tree_heading_bindings:
                try:
                    self.tree.heading(column_id, text=self._(key))
                except tk.TclError:
                    continue

        self._update_medium_options()
        self._update_help_text()

    def _resource_path(self, *parts: str) -> str:
        if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
            base = sys._MEIPASS
        else:
            base = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(base, "data", *parts)

    def _load_json_file(self, rel_path: str) -> typing.Any:
        try:
            with open(self._resource_path(rel_path), "r", encoding="utf-8") as f:
                return json.load(f)
        except OSError:
            return None
        except json.JSONDecodeError:
            logging.error("JSON parse error for %s", rel_path)
            return None

    def _load_external_resources(self) -> None:
        # Translations
        translations = self._load_json_file("translations.json")
        if isinstance(translations, dict):
            self.TRANSLATIONS = translations
        tooltips = self._load_json_file("tooltips.json")
        if isinstance(tooltips, dict):
            self.TOOLTIPS = tooltips
        # Help texts
        helps: dict[str, str] = {}
        for code in ("ru", "sr", "en"):
            try:
                with open(self._resource_path("help", f"{code}.txt"), "r", encoding="utf-8") as f:
                    helps[code] = f.read()
            except OSError:
                continue
        if helps:
            self.HELP_TEXTS = helps

        # Tables and numeric data
        tables = self._load_json_file("tables.json")
        if not isinstance(tables, dict):
            return
        self.INSULATION_OPTIONS = tables.get("INSULATION_OPTIONS", [])
        self.INSULATION_META = tables.get("INSULATION_META", {})
        self.CONDUCTOR_TYPES = tables.get("CONDUCTOR_TYPES", [])
        self.VOLTAGE_LEVELS = tables.get("VOLTAGE_LEVELS", [])
        self.TEMPERATURE_MEDIA = tables.get("TEMPERATURE_MEDIA", {})
        self.INSTALLATION_METHODS = tables.get("INSTALLATION_METHODS", [])
        self.STANDARD_SECTIONS = tables.get("STANDARD_SECTIONS", [])
        self.METHOD_PREFERENCE = tables.get("METHOD_PREFERENCE", [])
        self.STANDARD_CROSS_SECTIONS = tables.get("STANDARD_CROSS_SECTIONS", [])
        self.STANDARD_BREAKER_RATINGS = tables.get("STANDARD_BREAKER_RATINGS", [])
        self.DROP_LIMIT_KEYS = tables.get("DROP_LIMIT_KEYS", {})
        self.RESISTIVITY_20 = tables.get("RESISTIVITY_20", {})
        self.TEMP_COEFF = tables.get("TEMP_COEFF", {})
        self.REACTANCE_DATA = tables.get("REACTANCE", {})

        # Convert dicts that require numeric keys
        amp_base_raw = tables.get("AMPACITY_BASE", {})
        amp_base: dict[str, dict[float, float]] = {}
        for m, inner in amp_base_raw.items():
            try:
                amp_base[m] = {float(k): float(v) for k, v in inner.items()}
            except (ValueError, TypeError) as e:
                logging.warning("AMPACITY_BASE: skip method %s: %s", m, e)
                continue
        if amp_base:
            self.AMPACITY_BASE = amp_base

        self.AMPACITY_INSULATION_FACTORS = tables.get("AMPACITY_INSULATION_FACTORS", {})

        loaded_raw = tables.get("AMPACITY_LOADED_FACTORS", {})
        loaded: dict[str, dict[int, float]] = {}
        for m, inner in loaded_raw.items():
            try:
                loaded[m] = {int(k): float(v) for k, v in inner.items()}
            except (ValueError, TypeError) as e:
                logging.warning("AMPACITY_LOADED_FACTORS: skip method %s: %s", m, e)
                continue
        if loaded:
            self.AMPACITY_LOADED_FACTORS = loaded

        grouping_raw = tables.get("GROUPING_FACTORS", {})
        try:
            self.GROUPING_FACTORS = {int(k): float(v) for k, v in grouping_raw.items()}
        except (ValueError, TypeError) as e:
            logging.warning("GROUPING_FACTORS load failed: %s", e)

        ktv_raw = tables.get("KT_V_TABLE", {})
        ktv: dict[str, dict[int, float]] = {}
        for ins, inner in ktv_raw.items():
            try:
                ktv[ins] = {int(k): float(v) for k, v in inner.items()}
            except (ValueError, TypeError) as e:
                logging.warning("KT_V_TABLE: skip insulation %s: %s", ins, e)
                continue
        if ktv:
            self.KT_V_TABLE = ktv

        ktz_raw = tables.get("KT_Z_TABLE", {})
        ktz: dict[str, dict[int, float]] = {}
        for ins, inner in ktz_raw.items():
            try:
                ktz[ins] = {int(k): float(v) for k, v in inner.items()}
            except (ValueError, TypeError) as e:
                logging.warning("KT_Z_TABLE: skip insulation %s: %s", ins, e)
                continue
        if ktz:
            self.KT_Z_TABLE = ktz

    def _update_medium_options(self) -> None:
        if not self._medium_combobox:
            return
        language = self._language.get()
        values = [meta.get(language, meta.get(self.DEFAULT_LANGUAGE, "")) for meta in self.TEMPERATURE_MEDIA.values()]
        self._medium_combobox.configure(values=values)
        self._combobox_values["Среда для Т"] = values
        display = self.TEMPERATURE_MEDIA[self._medium_selected_key].get(
            language, self.TEMPERATURE_MEDIA[self._medium_selected_key][self.DEFAULT_LANGUAGE]
        )
        self._medium_combobox.set(display)
        var = self._form_values.get("Среда для Т")
        if var is not None:
            var.set(display)

    def _update_help_text(self) -> None:
        if self._help_text_widget is None:
            return
        language = self._language.get()
        help_text = self.HELP_TEXTS.get(language, self.HELP_TEXTS[self.DEFAULT_LANGUAGE])
        self._help_text_widget.configure(state="normal")
        self._help_text_widget.delete("1.0", tk.END)
        self._help_text_widget.insert("1.0", help_text)
        self._help_text_widget.configure(state="disabled")

    def _on_medium_changed(self, _: tk.Event | None) -> None:
        display = self._form_values.get("Среда для Т")
        if display is None:
            return
        selected = display.get().strip()
        for key, translations in self.TEMPERATURE_MEDIA.items():
            if selected in translations.values():
                self._medium_selected_key = key
                break
        self._update_intermediate_results()

    def _set_medium_from_value(self, value: str) -> None:
        normalized = value.strip()
        for key, translations in self.TEMPERATURE_MEDIA.items():
            if normalized in translations.values():
                self._medium_selected_key = key
                break
        language = self._language.get()
        display = self.TEMPERATURE_MEDIA[self._medium_selected_key].get(
            language, self.TEMPERATURE_MEDIA[self._medium_selected_key][self.DEFAULT_LANGUAGE]
        )
        medium_var = self._form_values.get("Среда для Т")
        if medium_var is not None:
            medium_var.set(display)
        if self._medium_combobox is not None:
            self._medium_combobox.set(display)

    def _attach_tooltip(self, widget: ttk.Widget, label_key: str) -> None:
        tooltip_texts = self.TOOLTIPS.get(label_key)
        if not tooltip_texts:
            return

        def text_getter(key: str = label_key) -> str:
            translations = self.TOOLTIPS.get(key, {})
            language = self._language.get()
            if language in translations:
                return translations[language]
            return translations.get(self.DEFAULT_LANGUAGE, "")

        tooltip = Tooltip(widget, text_getter)
        self._tooltips.append(tooltip)

    def _build_menu(self) -> None:
        menubar = tk.Menu(self)

        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label=self._("menu.save_project"), command=self.save_project)
        self._register_menu_text(file_menu, file_menu.index("end"), "menu.save_project")
        file_menu.add_command(label=self._("menu.load_project"), command=self.load_project)
        self._register_menu_text(file_menu, file_menu.index("end"), "menu.load_project")
        file_menu.add_separator()
        file_menu.add_command(label=self._("menu.export_excel"), command=self.export_to_excel)
        self._register_menu_text(file_menu, file_menu.index("end"), "menu.export_excel")

        menubar.add_cascade(label=self._("menu.file"), menu=file_menu)
        self._register_menu_text(menubar, menubar.index("end"), "menu.file")

        language_menu = tk.Menu(menubar, tearoff=0)
        for code, name in self.LANGUAGES.items():
            language_menu.add_radiobutton(
                label=name,
                variable=self._language,
                value=code,
                command=self._apply_language,
            )
        menubar.add_cascade(label=self._("menu.language"), menu=language_menu)
        self._register_menu_text(menubar, menubar.index("end"), "menu.language")

        self.config(menu=menubar)

    def _build_layout(self) -> None:
        notebook = ttk.Notebook(self)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        self._notebook = notebook

        main_tab = ttk.Frame(notebook)
        notebook.add(main_tab, text="")
        self._register_notebook_tab(main_tab, "tab.calculation")
        self._tabs["tab.calculation"] = main_tab

        help_tab = ttk.Frame(notebook)
        notebook.add(help_tab, text="")
        self._register_notebook_tab(help_tab, "tab.help")
        self._tabs["tab.help"] = help_tab

        container = ttk.Frame(main_tab)
        container.pack(fill=tk.BOTH, expand=True)

        form_frame = ttk.LabelFrame(container)
        form_frame.pack(fill=tk.X, expand=False, side=tk.TOP, pady=(0, 10))
        self._bind_text(lambda value, widget=form_frame: widget.configure(text=value), "frame.input")
        self._build_form(form_frame)

        intermediate_frame = ttk.LabelFrame(container)
        intermediate_frame.pack(fill=tk.X, expand=False, side=tk.TOP, pady=(0, 10))
        self._bind_text(
            lambda value, widget=intermediate_frame: widget.configure(text=value), "frame.intermediate"
        )
        self._build_intermediate_panel(intermediate_frame)

        table_frame = ttk.LabelFrame(container)
        table_frame.pack(fill=tk.BOTH, expand=True, side=tk.TOP)
        self._bind_text(lambda value, widget=table_frame: widget.configure(text=value), "frame.table")
        self._build_table(table_frame)

        self._register_form_traces()

        self._build_help_tab(help_tab)

    def _build_help_tab(self, parent: ttk.Frame) -> None:
        container = ttk.Frame(parent, padding=15)
        container.pack(fill=tk.BOTH, expand=True)
        container.columnconfigure(0, weight=1)
        container.rowconfigure(0, weight=1)

        scrollbar = ttk.Scrollbar(container)
        scrollbar.grid(row=0, column=1, sticky=tk.NS)

        font_name = "Segoe UI" if self.tk.call("tk", "windowingsystem") == "win32" else "TkDefaultFont"
        text = tk.Text(
            container,
            wrap="word",
            height=28,
            font=(font_name, 10),
            relief=tk.FLAT,
            state="disabled",
            cursor="arrow",
            yscrollcommand=scrollbar.set,
        )
        text.grid(row=0, column=0, sticky=tk.NSEW)
        scrollbar.configure(command=text.yview)
        self._help_text_widget = text
        self._update_help_text()

    def _build_form(self, parent: ttk.Frame) -> None:
        default_medium_display = self.TEMPERATURE_MEDIA[self._medium_selected_key][self.DEFAULT_LANGUAGE]
        field_specs = [
            ("Strujni krug", ""),
            ("Deonica OD", ""),
            ("Deonica DO", ""),
            ("Tip-IZOLACIJE", self.INSULATION_OPTIONS[0]),
            ("Tip-PROVODNIKA", self.CONDUCTOR_TYPES[0]),
            ("Oznaka-tip-KABLA", ""),
            ("Pi, W", ""),
            ("Kj", ""),
            ("η", "1.0"),
            ("U", self.VOLTAGE_LEVELS[0]),
            ("cos φ", ""),
            ("Dužina L, m", ""),
            ("Presek, mm²", ""),
            ("Način polaganja", self.DEFAULT_INSTALLATION_METHOD if self.DEFAULT_INSTALLATION_METHOD in self.INSTALLATION_METHODS else self.INSTALLATION_METHODS[0]),
            ("Нагруженные жилы (nž)", "3"),
            ("Кабелей в группе (для S)", "1"),
            ("Параллельные кабели (n∥)", "1"),
            ("Среда для Т", default_medium_display),
            ("Температура, °C", "30"),
            ("In, A", ""),
            ("k", "1.45"),
            ("Ключ ΔU", list(self.DROP_LIMIT_KEYS.keys())[0]),
        ]

        grid = ttk.Frame(parent)
        grid.pack(fill=tk.X, expand=False, padx=10, pady=10)

        columns = 4
        rows_per_column = math.ceil(len(field_specs) / columns)

        entry_columns = [col * 2 + 1 for col in range(columns)]
        for col in range(columns * 2):
            weight = 1 if col in entry_columns else 0
            grid.columnconfigure(col, weight=weight)

        for index, (label, default) in enumerate(field_specs):
            column = index // rows_per_column
            row = index % rows_per_column
            label_col = column * 2
            entry_col = label_col + 1

            label_widget = ttk.Label(grid, text="", anchor="e", justify="right")
            label_widget.grid(row=row, column=label_col, sticky=tk.E, pady=4, padx=(0, 8))
            label_key = self.LABEL_KEY_MAP.get(label, label)
            self._bind_text(lambda value, widget=label_widget: widget.configure(text=value), label_key)

            var = tk.StringVar(value=default)
            self._form_values[label] = var

            if label == "Tip-IZOLACIJE":
                widget = ttk.Combobox(grid, textvariable=var, values=self.INSULATION_OPTIONS, state="readonly")
            elif label == "Tip-PROVODNIKA":
                widget = ttk.Combobox(grid, textvariable=var, values=self.CONDUCTOR_TYPES, state="readonly")
            elif label == "U":
                widget = ttk.Combobox(grid, textvariable=var, values=self.VOLTAGE_LEVELS, state="readonly")
                self._combobox_values[label] = list(self.VOLTAGE_LEVELS)
            elif label == "Način polaganja":
                widget = ttk.Combobox(grid, textvariable=var, values=self.INSTALLATION_METHODS, state="readonly")
                self._combobox_values[label] = list(self.INSTALLATION_METHODS)
            elif label == "Нагруженные жилы (nž)":
                widget = ttk.Combobox(grid, textvariable=var, values=["2", "3"], state="readonly")
                self._combobox_values[label] = ["2", "3"]
            elif label == "Кабелей в группе (для S)":
                widget = ttk.Combobox(
                    grid, textvariable=var, values=[str(i) for i in range(1, 21)], state="readonly"
                )
                self._combobox_values[label] = [str(i) for i in range(1, 21)]
            elif label == "Параллельные кабели (n∥)":
                widget = ttk.Combobox(
                    grid, textvariable=var, values=[str(i) for i in range(1, 7)], state="readonly"
                )
                self._combobox_values[label] = [str(i) for i in range(1, 7)]
            elif label == "Среда для Т":
                medium_values = [
                    meta.get(self._language.get(), meta.get(self.DEFAULT_LANGUAGE, ""))
                    for meta in self.TEMPERATURE_MEDIA.values()
                ]
                widget = ttk.Combobox(grid, textvariable=var, values=medium_values, state="readonly")
                widget.bind("<<ComboboxSelected>>", self._on_medium_changed)
                self._medium_combobox = widget
                self._combobox_values[label] = medium_values
            elif label == "Ключ ΔU":
                widget = ttk.Combobox(
                    grid, textvariable=var, values=list(self.DROP_LIMIT_KEYS.keys()), state="readonly"
                )
                self._combobox_values[label] = list(self.DROP_LIMIT_KEYS.keys())
            elif label == "Температура, °C":
                widget = ttk.Entry(grid, textvariable=var)
                widget.bind("<FocusIn>", self._on_temperature_focus_in)
                widget.bind("<FocusOut>", self._on_temperature_focus_out)
            elif label == "Presek, mm²":
                widget = ttk.Combobox(
                    grid,
                    textvariable=var,
                    values=self.STANDARD_CROSS_SECTIONS,
                    state="readonly",
                )
                self._combobox_values[label] = list(self.STANDARD_CROSS_SECTIONS)
            elif label == "In, A":
                widget = ttk.Combobox(
                    grid,
                    textvariable=var,
                    values=self.STANDARD_BREAKER_RATINGS,
                    state="readonly",
                )
                self._combobox_values[label] = list(self.STANDARD_BREAKER_RATINGS)
            else:
                widget = ttk.Entry(grid, textvariable=var)

            widget.grid(row=row, column=entry_col, sticky=tk.EW, pady=4)
            widget_class = widget.winfo_class()
            original_style = widget.cget("style") or widget_class
            self._input_widgets[label] = widget
            self._input_styles[label] = original_style
            self._attach_tooltip(widget, label_key)

        last_field_row_in_last_col = (len(field_specs) - 1) % rows_per_column
        select_button = ttk.Button(grid, text="", command=self.select_optimal_parameters)

        select_button.grid(row=last_field_row_in_last_col + 2, column=7, columnspan=1, padx=10, pady=0)

        self._bind_text(
            lambda value, widget=select_button: widget.configure(text=value), "button.select_optimal"
        )

        pi_var = self._form_values["Pi, W"]
        kj_var = self._form_values["Kj"]
        pi_var.trace_add("write", self._update_pj_display)
        kj_var.trace_add("write", self._update_pj_display)

        for key, default in (("Pj", ""), ("S", "1.0"), ("T", "1.0")):
            if key not in self._form_values:
                self._form_values[key] = tk.StringVar(value=default)

    def _build_intermediate_panel(self, parent: ttk.Frame) -> None:
        grid = ttk.Frame(parent)
        grid.pack(fill=tk.X, expand=False, padx=10, pady=10)

        specs = [
            ("Pj, W", "Pj, W"),
            ("Icalc [A]", "Icalc [A]"),
            ("R_base [Ω/km]", "R_base [Ω/km]"),
            ("Iz [A]", "Iz [A]"),
            ("S", "S"),
            ("T", "T"),
            ("ΔU %", "ΔU %"),
            ("Ukupni ΔU %", "Ukupni ΔU %"),
            ("Limit ΔU %", "Limit ΔU %"),
            ("По току", "По току"),
            ("По ΔU", "По ΔU"),
            ("Диапазон In [A]", "Диапазон In [A]"),
            ("I2 [A]", "I2 [A]"),
            ("Защита", "Защита"),
            ("Совместимость IEC", "Совместимость IEC"),
            ("Рекомендации", "Рекомендации"),
        ]

        columns = 3
        for col in range(columns * 2):
            weight = 1 if col % 2 == 1 else 0
            grid.columnconfigure(col, weight=weight)

        for index, (label_text, key) in enumerate(specs):
            row = index // columns
            label_col = (index % columns) * 2
            value_col = label_col + 1

            label_widget = ttk.Label(grid, text="", style="ResultKey.TLabel")
            label_widget.grid(row=row, column=label_col, sticky=tk.W, pady=4, padx=(0, 8))
            label_key = self.RESULT_LABEL_KEY_MAP.get(label_text, label_text)
            self._bind_text(lambda value, widget=label_widget: widget.configure(text=value), label_key)

            var = tk.StringVar(value="—")
            value_label = ttk.Label(grid, textvariable=var, style="ResultValue.TLabel")
            value_label.grid(row=row, column=value_col, sticky=tk.EW, pady=4)
            self._intermediate_vars[key] = var
            self._intermediate_labels[key] = value_label

        button_frame = ttk.Frame(parent)
        button_frame.pack(fill=tk.X, pady=(10, 0))

        add_button = ttk.Button(button_frame, text="", command=self.add_row)
        add_button.pack(side=tk.LEFT, padx=(0, 5))
        self._bind_text(lambda value, widget=add_button: widget.configure(text=value), "button.add_row")

        load_button = ttk.Button(button_frame, text="", command=self.load_selected_row)
        load_button.pack(side=tk.LEFT, padx=(0, 5))
        self._bind_text(lambda value, widget=load_button: widget.configure(text=value), "button.load_row")

        delete_button = ttk.Button(button_frame, text="", command=self.remove_selected_row)
        delete_button.pack(side=tk.LEFT, padx=(0, 5))
        self._bind_text(lambda value, widget=delete_button: widget.configure(text=value), "button.remove_row")

        clear_button = ttk.Button(button_frame, text="", command=self.clear_table)
        clear_button.pack(side=tk.LEFT, padx=(0, 5))
        self._bind_text(lambda value, widget=clear_button: widget.configure(text=value), "button.clear")

        export_button = ttk.Button(button_frame, text="", command=self.export_to_excel)
        export_button.pack(side=tk.RIGHT)
        self._bind_text(lambda value, widget=export_button: widget.configure(text=value), "menu.export_excel")

    def _build_table(self, parent: ttk.Frame) -> None:
        columns = self.TREE_COLUMNS

        tree_container = ttk.Frame(parent)
        tree_container.pack(fill=tk.BOTH, expand=True)
        tree_container.columnconfigure(0, weight=1)
        tree_container.rowconfigure(0, weight=1)

        tree = ttk.Treeview(tree_container, columns=columns, show="headings")
        self.tree = tree

        for col in columns:
            key = self.TREE_COLUMN_KEYS.get(col, col)
            self._register_tree_heading(col, key)
            tree.column(col, width=120, anchor=tk.CENTER)

        tree.grid(row=0, column=0, sticky=tk.NSEW)

        v_scroll = ttk.Scrollbar(tree_container, orient=tk.VERTICAL, command=tree.yview)
        v_scroll.grid(row=0, column=1, sticky=tk.NS)

        h_scroll = ttk.Scrollbar(tree_container, orient=tk.HORIZONTAL, command=tree.xview)
        h_scroll.grid(row=1, column=0, sticky=tk.EW)

        tree.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)
        tree.bind("<Double-1>", lambda e: self.load_selected_row())

    def _register_form_traces(self) -> None:
        for name, var in self._form_values.items():
            if name in {"Pj", "S", "T"}:
                continue
            var.trace_add("write", self._update_intermediate_results)
        self._update_intermediate_results()

    def _try_parse_float(self, value: str) -> float | None:
        value = value.strip().replace(",", ".")
        if not value:
            return None
        try:
            return float(value)
        except ValueError:
            return None

    def _parse_float(self, value: str, field_name: str) -> float | None:
        value = value.strip()
        field_display = self._(self.LABEL_KEY_MAP.get(field_name, field_name))
        if not value:
            messagebox.showerror(self._("title.error_input"), self._m("error.input_empty", field_display))
            logging.error("Пустое значение в поле '%s'", field_name)
            return None
        try:
            return float(value.replace(",", "."))
        except ValueError:
            messagebox.showerror(self._("title.error_input"), self._m("error.input_invalid", field_display, value))
            logging.error("Некорректное значение '%s' в поле '%s'", value, field_name)
            return None

    def _fmt(self, value: float | None, digits: int = 2) -> str:
        if value is None:
            return "—"
        try:
            if not math.isfinite(value):
                return "—"
        except TypeError:
            return "—"
        formatted = f"{value:.{digits}f}"
        if self._language.get() in {"ru", "sr"}:
            formatted = formatted.replace(".", ",")
        return formatted

    def _lookup_ampacity(
        self, insulation_key: str, conductor: str, laying: str, area: float, loaded_cores: int
    ) -> float | None:
        base_table = self.AMPACITY_BASE.get(laying)
        if not base_table:
            return None

        insulation_factors = self.AMPACITY_INSULATION_FACTORS.get(insulation_key, {}).get(conductor)
        load_factors = self.AMPACITY_LOADED_FACTORS.get(laying)
        if insulation_factors is None or load_factors is None:
            return None

        multiplier = insulation_factors.get(laying)
        load_multiplier = load_factors.get(loaded_cores)
        if multiplier is None or load_multiplier is None:
            return None

        standard_sections = sorted(base_table)
        if not standard_sections:
            return None

        if area <= standard_sections[0]:
            base_value = base_table[standard_sections[0]]
        elif area >= standard_sections[-1]:
            base_value = base_table[standard_sections[-1]]
        else:
            for lower, upper in zip(standard_sections, standard_sections[1:]):
                if math.isclose(area, lower, rel_tol=1e-6, abs_tol=1e-3):
                    base_value = base_table[lower]
                    break
                if lower <= area <= upper:
                    lower_val = base_table[lower]
                    upper_val = base_table[upper]
                    if math.isclose(upper, lower):
                        base_value = lower_val
                    else:
                        ratio = (area - lower) / (upper - lower)
                        base_value = lower_val + ratio * (upper_val - lower_val)
                    break
            else:
                return None

        return base_value * multiplier * load_multiplier

    def _lookup_group_factor(self, circuits: int) -> float:
        if circuits <= 1:
            return 1.0
        max_defined = max(self.GROUPING_FACTORS)
        if circuits >= max_defined:
            return self.GROUPING_FACTORS[max_defined]
        return self.GROUPING_FACTORS.get(circuits, self.GROUPING_FACTORS[max_defined])

    def _lookup_temperature_factor(self, insulation_key: str, medium: str, temperature: float) -> float | None:
        if medium == "soil":
            table = self.KT_Z_TABLE.get(insulation_key)
        else:
            table = self.KT_V_TABLE.get(insulation_key)
        if not table:
            return None

        points = sorted(table)
        if not points:
            return None
        if temperature < points[0] or temperature > points[-1]:
            return None

        if temperature in table:
            return table[temperature]

        for lower, upper in zip(points, points[1:]):
            if lower <= temperature <= upper:
                lower_val = table[lower]
                upper_val = table[upper]
                if math.isclose(upper, lower):
                    return lower_val
                ratio = (temperature - lower) / (upper - lower)
                return lower_val + ratio * (upper_val - lower_val)
        return None

    def _calculate_line_impedance(
        self, conductor: str, insulation_temp: float, area: float, laying: str
    ) -> tuple[float, float]:
        rho_20 = self.RESISTIVITY_20.get(conductor)
        alpha = self.TEMP_COEFF.get(conductor)
        if rho_20 is None or alpha is None:
            default_x = 0.08
            if isinstance(getattr(self, "REACTANCE_DATA", None), dict) and self.REACTANCE_DATA:
                default_x = self.REACTANCE_DATA.get("default", default_x)
            return 0.0, default_x

        rho_theta = rho_20 * (1.0 + alpha * (insulation_temp - 20.0))
        if area <= 0:
            r_per_km = 0.0
        else:
            r_per_km = (rho_theta / area) * 1000.0

        bucket = "≤95"
        if area > 240:
            bucket = ">240"
        elif area > 95:
            bucket = "≤240"
        # Prefer externally loaded reactance data if available
        x_per_km = None
        if isinstance(getattr(self, "REACTANCE_DATA", None), dict) and self.REACTANCE_DATA:
            try:
                buckets = self.REACTANCE_DATA.get("buckets", {})
                method_defaults = self.REACTANCE_DATA.get("method_defaults", {})
                x_per_km = buckets.get(laying, {}).get(bucket)
                if x_per_km is None:
                    x_per_km = method_defaults.get(laying)
                if x_per_km is None:
                    x_per_km = self.REACTANCE_DATA.get("default")
            except (KeyError, TypeError, ValueError) as e:
                logging.debug("REACTANCE_DATA lookup failed: %s", e)
                x_per_km = None
        if x_per_km is None:
            if isinstance(getattr(self, "REACTANCE_DATA", None), dict):
                x_per_km = self.REACTANCE_DATA.get("default", 0.08)
            else:
                x_per_km = 0.08
        return r_per_km, x_per_km

    def _drop_pct(
        self,
        U: float,
        cos_phi: float,
        L_m: float,
        conductor: str,
        insulation_theta: float,
        area_mm2: float,
        method: str,
        loaded_cores: int,
        n_parallel: int,
        icalc_total: float,
    ) -> float:
        r_km, x_km = self._calculate_line_impedance(conductor, insulation_theta, area_mm2, method)
        divider = max(n_parallel, 1)
        r_m, x_m = (r_km / 1000.0) / divider, (x_km / 1000.0) / divider
        sin_phi = math.sqrt(max(0.0, 1.0 - min(1.0, cos_phi) ** 2))
        phase = 2.0 if loaded_cores == 2 else math.sqrt(3)
        if U == 0:
            return 0.0
        return phase * icalc_total * (r_m * cos_phi + x_m * sin_phi) * L_m * 100.0 / U

    def _recommend(
        self,
        *,
        U: float,
        cos_phi: float,
        L: float,
        conductor: str,
        insulation_key: str,
        insulation_theta: float,
        method: str,
        loaded_cores: int,
        S: float,
        T: float,
        limit_pct: float | None,
        current_area: float,
        n_parallel: int,
        icalc_total: float,
    ) -> list[str]:
        recs: list[str] = []
        if not icalc_total or S <= 0 or T <= 0:
            return recs

        parallel_count = max(n_parallel, 1)
        current_per_cable = icalc_total / parallel_count if parallel_count else icalc_total
        min_section = max(current_area or 0, self.STANDARD_SECTIONS[0])

        def fmt(value: float, digits: int = 2) -> str:
            return self._fmt(value, digits=digits)

        def candidate_description(action: str, area_val: float, method_val: str, iz_one: float, drop_val: float) -> str:
            digits = 0 if float(area_val).is_integer() else 1
            area_str = fmt(area_val, digits=digits)
            iz_total = iz_one * parallel_count
            return (
                f"{action} {area_str} мм² (метод {method_val}) → "
                f"Iz_tot≈{fmt(iz_total, digits=0)} A, ΔU≈{fmt(drop_val)}%"
            )

        # 1) Increase section within same method/insulation
        best_section_line = None
        for area_candidate in self.STANDARD_SECTIONS:
            if area_candidate < min_section:
                continue
            iz_base = self._lookup_ampacity(insulation_key, conductor, method, area_candidate, loaded_cores)
            if not iz_base:
                continue
            iz_one = iz_base * S * T
            if iz_one < current_per_cable:
                continue
            drop_val = self._drop_pct(
                U,
                cos_phi,
                L,
                conductor,
                insulation_theta,
                area_candidate,
                method,
                loaded_cores,
                parallel_count,
                icalc_total,
            )
            if limit_pct is not None and drop_val > limit_pct:
                continue
            best_section_line = candidate_description("Увеличить сечение до", area_candidate, method, iz_one, drop_val)
            break
        if best_section_line:
            recs.append(best_section_line)

        # 2) Change method according to preference
        best_method_line = None
        for m in self.METHOD_PREFERENCE:
            if m == method:
                continue
            for area_candidate in self.STANDARD_SECTIONS:
                if area_candidate < min_section:
                    continue
                iz_base = self._lookup_ampacity(insulation_key, conductor, m, area_candidate, loaded_cores)
                if not iz_base:
                    continue
                iz_one = iz_base * S * T
                if iz_one < current_per_cable:
                    continue
                drop_val = self._drop_pct(
                    U,
                    cos_phi,
                    L,
                    conductor,
                    insulation_theta,
                    area_candidate,
                    m,
                    loaded_cores,
                    parallel_count,
                    icalc_total,
                )
                if limit_pct is not None and drop_val > limit_pct:
                    continue
                best_method_line = candidate_description("Сменить метод на", area_candidate, m, iz_one, drop_val)
                break
            if best_method_line:
                break
        if best_method_line:
            recs.append(best_method_line)

        # 3) Switch to XLPE if currently PVC
        if insulation_key == "PVC":
            xlpe_meta = self.INSULATION_META.get("XLPE/EPR (90°C)")
            xlpe_theta = xlpe_meta.get("theta", 90.0) if xlpe_meta else 90.0
            best_xlpe_line = None
            for area_candidate in self.STANDARD_SECTIONS:
                if area_candidate < min_section:
                    continue
                iz_base = self._lookup_ampacity("XLPE", conductor, method, area_candidate, loaded_cores)
                if not iz_base:
                    continue
                iz_one = iz_base * S * T
                if iz_one < current_per_cable:
                    continue
                drop_val = self._drop_pct(
                    U,
                    cos_phi,
                    L,
                    conductor,
                    xlpe_theta,
                    area_candidate,
                    method,
                    loaded_cores,
                    parallel_count,
                    icalc_total,
                )
                if limit_pct is not None and drop_val > limit_pct:
                    continue
                digits = 0 if float(area_candidate).is_integer() else 1
                area_str = fmt(area_candidate, digits=digits)
                iz_total = iz_one * parallel_count
                best_xlpe_line = (
                    f"Перейти на XLPE и {area_str} мм² (метод {method}) → "
                    f"Iz_tot≈{fmt(iz_total, digits=0)} A, ΔU≈{fmt(drop_val)}%"
                )
                break
            if best_xlpe_line:
                recs.append(best_xlpe_line)

        # 4) Increase number of parallel cables to meet voltage drop
        if limit_pct is not None:
            base_area = max(min_section, current_area or self.STANDARD_SECTIONS[0])
            drop_current = self._drop_pct(
                U,
                cos_phi,
                L,
                conductor,
                insulation_theta,
                base_area,
                method,
                loaded_cores,
                parallel_count,
                icalc_total,
            )
            if drop_current > limit_pct:
                iz_base = self._lookup_ampacity(insulation_key, conductor, method, base_area, loaded_cores)
                if iz_base:
                    iz_one = iz_base * S * T
                    denom = max(limit_pct, 1e-9)
                    n_needed = max(parallel_count + 1, math.ceil(drop_current * parallel_count / denom))
                    if iz_one > 0:
                        max_total = iz_one * n_needed
                        drop_new = self._drop_pct(
                            U,
                            cos_phi,
                            L,
                            conductor,
                            insulation_theta,
                            base_area,
                            method,
                            loaded_cores,
                            n_needed,
                            icalc_total,
                        )
                        digits = 0 if float(base_area).is_integer() else 1
                        area_str = fmt(base_area, digits=digits)
                        recs.append(
                            f"Разделить на {n_needed} параллельных кабеля {area_str} мм² (метод {method}) → "
                            f"n∥={n_needed}, Iz_tot≈{fmt(max_total, digits=0)} A, ΔU≈{fmt(drop_new)}%"
                        )

        if not recs:
            recs.append("Снизить число кабелей в группе (для увеличения S) или повысить напряжение.")
        return recs[:4]

    def _update_pj_display(self, *_: object) -> None:
        pi = self._try_parse_float(self._form_values["Pi, W"].get())
        kj = self._try_parse_float(self._form_values["Kj"].get())
        if pi is None or kj is None:
            self._form_values["Pj"].set("")
            self._update_intermediate_results()
            return
        self._form_values["Pj"].set(self._fmt(pi * kj))
        self._update_intermediate_results()

    def _set_result_alert(self, key: str, alert: bool) -> None:
        label = self._intermediate_labels.get(key)
        if not label:
            return
        label.configure(style="ResultAlert.TLabel" if alert else "ResultValue.TLabel")

    def _validate_counts(
        self,
    ) -> tuple[int, int, int, str | None]:
        """Parse loaded cores, circuits count, n_parallel from form. Returns (loaded_cores, circuits_count, n_parallel, error_key or None)."""
        try:
            loaded_cores = int(self._form_values["Нагруженные жилы (nž)"].get().strip())
            if loaded_cores not in (2, 3):
                return 3, 1, 1, "error.loaded_cores_2_3"
        except ValueError:
            return 3, 1, 1, "error.loaded_cores_2_3"
        try:
            circuits_count = int(self._form_values["Кабелей в группе (для S)"].get().strip())
            if circuits_count < 1:
                return loaded_cores, 1, 1, "error.circuits_min_1"
        except ValueError:
            return loaded_cores, 1, 1, "error.circuits_min_1"
        try:
            n_parallel = int(self._form_values["Параллельные кабели (n∥)"].get().strip())
            if n_parallel < 1:
                return loaded_cores, circuits_count, 1, "error.parallel_min_1"
        except ValueError:
            return loaded_cores, circuits_count, 1, "error.parallel_min_1"
        return loaded_cores, circuits_count, n_parallel, None

    def _set_entry_alert(self, field_name: str, alert: bool) -> None:
        widget = self._input_widgets.get(field_name)
        if widget is None:
            return
        default_style = self._input_styles.get(field_name, "")
        if not alert:
            widget.configure(style=default_style)
            return
        widget_class = widget.winfo_class()
        if widget_class == "TEntry":
            widget.configure(style="Invalid.TEntry")
        elif widget_class == "TCombobox":
            widget.configure(style="Invalid.TCombobox")

    def _on_temperature_focus_in(self, _: tk.Event) -> None:
        self._temperature_editing = True

    def _on_temperature_focus_out(self, _: tk.Event) -> None:
        self._temperature_editing = False
        self._update_intermediate_results()

    def _show_temperature_warning(self, insulation_key: str, medium: str, temperature: float) -> None:
        if self._temperature_editing:
            widget = self._input_widgets.get("Температура, °C")
            if widget is not None and self.focus_get() == widget:
                return
        rounded_temp = round(temperature, 1)
        key = (insulation_key, medium, rounded_temp)
        if self._last_temperature_warning == key:
            return
        self._last_temperature_warning = key
        logging.warning(
            "Температура вне диапазона для изоляции %s, среды %s: %s °C",
            insulation_key,
            medium,
            rounded_temp,
        )
        messagebox.showwarning(
            self._("title.temperature_range"),
            self._m("warning.temperature_range_msg", rounded_temp),
        )

    def _update_intermediate_results(self, *_: object) -> None:
        if not self._intermediate_vars:
            return
        if self._loading_project or self._loading_row_into_form:
            return

        self._last_icalc = None
        for key, var in self._intermediate_vars.items():
            if key == "Limit ΔU %":
                continue
            var.set("—")

        for key in ("ΔU %", "По ΔU", "По току", "Защита", "Ukupni ΔU %"):
            self._set_result_alert(key, False)

        strujni_krug = self._form_values["Strujni krug"].get().strip()
        insulation_label = self._form_values["Tip-IZOLACIJE"].get()
        insulation_meta = self.INSULATION_META.get(insulation_label)
        conductor = self._form_values["Tip-PROVODNIKA"].get()
        laying = self._form_values["Način polaganja"].get().strip()
        loaded_cores_value = self._form_values["Нагруженные жилы (nž)"].get().strip()
        group_value = self._form_values["Кабелей в группе (для S)"].get().strip()
        parallel_value = self._form_values["Параллельные кабели (n∥)"].get().strip()
        if self._medium_selected_key not in self.TEMPERATURE_MEDIA:
            self._medium_selected_key = next(iter(self.TEMPERATURE_MEDIA))
            self._update_medium_options()
        medium_key = self._medium_selected_key
        drop_key = self._form_values["Ключ ΔU"].get()
        limit_delta = self.DROP_LIMIT_KEYS.get(drop_key)
        if limit_delta is not None:
            self._intermediate_vars["Limit ΔU %"].set(self._fmt(limit_delta))
        else:
            self._intermediate_vars["Limit ΔU %"].set("—")

        loaded_cores = 3
        cores_alert = False
        try:
            loaded_cores = int(loaded_cores_value)
            if loaded_cores not in (2, 3):
                raise ValueError
        except ValueError:
            cores_alert = True
            loaded_cores = 3
        self._set_entry_alert("Нагруженные жилы (nž)", cores_alert)

        circuits_alert = False
        try:
            circuits_count = int(group_value)
            if circuits_count < 1:
                raise ValueError
        except ValueError:
            circuits_count = 1
            circuits_alert = True
        self._set_entry_alert("Кабелей в группе (для S)", circuits_alert)

        parallel_alert = False
        try:
            n_parallel = int(parallel_value)
            if n_parallel < 1:
                raise ValueError
        except ValueError:
            n_parallel = 1
            parallel_alert = True
        self._set_entry_alert("Параллельные кабели (n∥)", parallel_alert)

        pi = self._try_parse_float(self._form_values["Pi, W"].get())
        kj = self._try_parse_float(self._form_values["Kj"].get())
        eta_value = self._try_parse_float(self._form_values["η"].get())
        cos_phi = self._try_parse_float(self._form_values["cos φ"].get())
        length = self._try_parse_float(self._form_values["Dužina L, m"].get())
        area = self._try_parse_float(self._form_values["Presek, mm²"].get())
        temperature = self._try_parse_float(self._form_values["Температура, °C"].get())

        effective_circuits = max(1, circuits_count + n_parallel - 1) if n_parallel > 1 else circuits_count
        group_factor = self._lookup_group_factor(effective_circuits)
        s_display = self._fmt(group_factor)
        self._form_values["S"].set(s_display if s_display != "—" else "")
        self._intermediate_vars["S"].set(s_display)
        s_coeff = group_factor

        temperature_alert = False
        t_coeff: float | None = None
        if insulation_meta is not None and temperature is not None:
            temp_factor = self._lookup_temperature_factor(insulation_meta["key"], medium_key, temperature)
            if temp_factor is not None and temp_factor > 0:
                t_coeff = temp_factor
            else:
                temperature_alert = True
        elif insulation_meta is not None:
            temperature_alert = True

        if t_coeff is not None:
            t_display = self._fmt(t_coeff)
            self._last_temperature_warning = None
        else:
            t_display = ""
            if (
                temperature is not None
                and insulation_meta is not None
                and not self._temperature_editing
            ):
                self._show_temperature_warning(insulation_meta["key"], medium_key, temperature)
        if t_display:
            self._form_values["T"].set(t_display)
            self._intermediate_vars["T"].set(t_display)
        else:
            self._form_values["T"].set("")
            self._intermediate_vars["T"].set("—")
        if self._temperature_editing:
            temperature_alert = False
        self._set_entry_alert("Температура, °C", temperature_alert)

        eta_alert = False
        eta_coeff: float | None = None
        if eta_value is not None:
            if eta_value <= 0 or eta_value > 1:
                eta_alert = True
            else:
                eta_coeff = eta_value
        else:
            eta_alert = True
        self._set_entry_alert("η", eta_alert)

        voltage_str = self._form_values["U"].get()
        try:
            voltage_value = int(voltage_str)
        except (TypeError, ValueError):
            voltage_value = None

        voltage_alert = voltage_value == 230 and loaded_cores == 3
        if voltage_alert and not self._voltage_phase_warning_shown and not self._loading_project:
            messagebox.showwarning(self._("title.iec"), self._("warning.voltage_phase"))
            self._voltage_phase_warning_shown = True
        self._set_entry_alert("U", voltage_alert)

        pj = None
        if pi is not None and kj is not None:
            pj = pi * kj
            self._intermediate_vars["Pj, W"].set(self._fmt(pj))

        cos_alert = False
        if cos_phi is not None:
            if not (0 < cos_phi <= 1):
                cos_alert = True
                cos_phi = None
        self._set_entry_alert("cos φ", cos_alert)

        area_alert = False
        if area is not None and area <= 0:
            area_alert = True
            area = None

        length_alert = False
        if length is not None and length < 0:
            length_alert = True
            length = None

        self._set_entry_alert("Presek, mm²", area_alert)
        self._set_entry_alert("Dužina L, m", length_alert)

        in_range_alert = False
        icalc_total = None
        icalc_per_cable = None
        phase_factor = None
        if pj is not None and cos_phi is not None and voltage_value and eta_coeff is not None:
            phase_factor = 2.0 if loaded_cores == 2 else math.sqrt(3)
            denominator = phase_factor * voltage_value * cos_phi
            if denominator:
                icalc_total = (pj / eta_coeff) / denominator
                self._intermediate_vars["Icalc [A]"].set(self._fmt(icalc_total, digits=3))
                self._last_icalc = icalc_total
                icalc_per_cable = icalc_total / max(n_parallel, 1)
        else:
            self._intermediate_vars["Icalc [A]"].set("—")

        r_per_km = None
        x_per_km = None
        if area is not None and insulation_meta is not None:
            r_per_km, x_per_km = self._calculate_line_impedance(
                conductor, insulation_meta["theta"], area, laying
            )
            self._intermediate_vars["R_base [Ω/km]"].set(self._fmt(r_per_km, digits=3))

        base_ampacity = None
        if area is not None and insulation_meta is not None:
            base_ampacity = self._lookup_ampacity(
                insulation_meta["key"], conductor, laying, area, loaded_cores
            )

        iz_one = None
        if base_ampacity is not None and t_coeff is not None:
            iz_one = base_ampacity * s_coeff * t_coeff
            self._intermediate_vars["Iz [A]"].set(self._fmt(iz_one))
        elif base_ampacity is None:
            self._intermediate_vars["Iz [A]"].set("—")

        in_range_value = "—"
        if icalc_total is not None and iz_one is not None:
            iz_total = iz_one * n_parallel
            in_range_value = f"{self._fmt(icalc_total)} – {self._fmt(iz_total)}"
            if iz_total + 1e-9 < icalc_total:
                in_range_alert = True
        if "Диапазон In [A]" in self._intermediate_vars:
            self._intermediate_vars["Диапазон In [A]"].set(in_range_value)

        ampacity_status = None
        if base_ampacity is None:
            ampacity_status = "N/A"
        elif iz_one is None or icalc_per_cable is None:
            ampacity_status = "—"
        else:
            ampacity_status = "OK" if icalc_per_cable <= iz_one else "NE"
        ampacity_alert = ampacity_status == "NE" or area_alert
        if ampacity_status is not None:
            self._intermediate_vars["По току"].set(ampacity_status)
            self._set_result_alert("По току", ampacity_alert)
        self._set_entry_alert("Presek, mm²", ampacity_alert)

        delta_u = None
        if (
            icalc_total is not None
            and phase_factor is not None
            and length is not None
            and cos_phi is not None
            and r_per_km is not None
            and x_per_km is not None
            and voltage_value
        ):
            divider = max(n_parallel, 1)
            r_per_meter = (r_per_km / 1000.0) / divider
            x_per_meter = (x_per_km / 1000.0) / divider
            sin_phi = math.sqrt(max(0.0, 1.0 - min(1.0, cos_phi) ** 2))
            impedance_drop = r_per_meter * cos_phi + x_per_meter * sin_phi
            delta_u = phase_factor * icalc_total * impedance_drop * length * 100.0 / voltage_value
            self._intermediate_vars["ΔU %"].set(self._fmt(delta_u))
        elif "ΔU %" in self._intermediate_vars:
            self._intermediate_vars["ΔU %"].set("—")

        drop_status = None
        if delta_u is not None and limit_delta is not None:
            drop_status = "OK" if delta_u <= limit_delta else "NE"
            self._intermediate_vars["По ΔU"].set(drop_status)
            self._set_result_alert("По ΔU", drop_status == "NE")
            self._set_result_alert("ΔU %", drop_status == "NE")
        elif delta_u is not None:
            drop_status = "—"
            self._intermediate_vars["По ΔU"].set(drop_status)

        drop_alert = (drop_status == "NE") or length_alert
        self._set_entry_alert("Dužina L, m", drop_alert)

        form_od = (self._form_values.get("Deonica OD") or tk.StringVar()).get().strip()
        existing_drop = self._sum_drop_chain_ending_at(strujni_krug, form_od)
        if delta_u is not None:
            total_drop = existing_drop + delta_u
            self._intermediate_vars["Ukupni ΔU %"].set(self._fmt(total_drop))
            if limit_delta is not None:
                total_status = "OK" if total_drop <= limit_delta else "NE"
                self._set_result_alert("Ukupni ΔU %", total_status == "NE")
            else:
                self._set_result_alert("Ukupni ΔU %", False)
        elif existing_drop > 0:
            self._intermediate_vars["Ukupni ΔU %"].set(self._fmt(existing_drop))

        in_value = self._try_parse_float(self._form_values["In, A"].get())
        k_value = self._try_parse_float(self._form_values["k"].get())
        i2_value = None
        protection_status = "—"
        protection_alert = False

        if in_value is not None and in_value <= 0:
            protection_alert = True
            in_value = None

        if k_value is not None and k_value <= 0:
            protection_alert = True
            k_value = None

        if in_value is not None and k_value is not None:
            i2_value = in_value * k_value
            self._intermediate_vars["I2 [A]"].set(self._fmt(i2_value))
        else:
            self._intermediate_vars["I2 [A]"].set("—")

        if iz_one is None or icalc_total is None or in_value is None or i2_value is None:
            if in_value is None or i2_value is None:
                protection_status = "—"
            else:
                protection_status = "N/A"
        else:
            iz_total = iz_one * n_parallel
            within_nominal = icalc_total <= in_value <= iz_total
            overload_check = i2_value <= 1.45 * iz_total
            if within_nominal and overload_check:
                protection_status = "OK"
            else:
                protection_status = "NE"
                protection_alert = True

        self._intermediate_vars["Защита"].set(protection_status)
        self._set_result_alert("Защита", protection_alert)
        self._set_result_alert("Диапазон In [A]", in_range_alert or protection_alert)
        self._set_entry_alert("In, A", protection_alert)

        if "Совместимость IEC" in self._intermediate_vars:
            if base_ampacity is None:
                compat_display = self._("status.na")
                compat_alert = False
            else:
                statuses = [ampacity_status, drop_status, protection_status]
                if all(status == "OK" for status in statuses):
                    compat_display = self._("status.ok")
                    compat_alert = False
                elif any(status == "NE" for status in statuses):
                    compat_display = self._("status.fail")
                    compat_alert = True
                else:
                    compat_display = self._("status.na")
                    compat_alert = False
            self._intermediate_vars["Совместимость IEC"].set(compat_display)
            self._set_result_alert("Совместимость IEC", compat_alert)

        if (
            icalc_total is not None
            and delta_u is not None
            and insulation_meta is not None
            and pj is not None
        ):
            self._last_result = {
                "pj": pj,
                "s_coeff": s_coeff,
                "t_coeff": t_coeff or 1.0,
                "icalc_total": icalc_total,
                "r_per_km": r_per_km,
                "sigma": (
                    1.0 / self.RESISTIVITY_20[conductor]
                    if conductor in self.RESISTIVITY_20 and self.RESISTIVITY_20[conductor] > 0
                    else None
                ),
                "iz_one": iz_one,
                "delta_u": delta_u,
                "ampacity_ok": ampacity_status,
                "drop_ok": drop_status,
                "protection_status": protection_status,
                "i2_value": i2_value,
                "in_value": in_value,
                "k_value": k_value,
                "compatibility_status": (
                    self._("status.no_data")
                    if base_ampacity is None
                    else self._("status.ok")
                    if (ampacity_status == "OK" and drop_status == "OK" and protection_status == "OK")
                    else self._("status.fail")
                    if (ampacity_status == "NE" or drop_status == "NE" or protection_status == "NE")
                    else self._("status.na")
                ),
            }
        else:
            self._last_result = None

        recommendations: list[str] = []
        rec_var = self._intermediate_vars.get("Рекомендации")
        if rec_var is not None:
            ampacity_state = self._intermediate_vars.get("По току")
            drop_state = self._intermediate_vars.get("По ΔU")
            ampacity_value = ampacity_state.get() if ampacity_state is not None else ""
            drop_value = drop_state.get() if drop_state is not None else ""
            if ampacity_value == "NE" or drop_value == "NE":
                insulation_label = self._form_values["Tip-IZOLACIJE"].get()
                insulation_meta = self.INSULATION_META.get(
                    insulation_label, {"key": "PVC", "theta": 70}
                )
                try:
                    loaded_cores_value = int(
                        self._form_values["Нагруженные жилы (nž)"].get() or "3"
                    )
                except (TypeError, ValueError):
                    loaded_cores_value = 3
                s_value = self._try_parse_float(self._form_values["S"].get())
                t_value = self._try_parse_float(self._form_values["T"].get())
                recs = self._recommend(
                    U=float(voltage_value or 0),
                    cos_phi=float(cos_phi or 0),
                    L=float(length or 0),
                    conductor=self._form_values["Tip-PROVODNIKA"].get(),
                    insulation_key=insulation_meta.get("key", "PVC"),
                    insulation_theta=float(insulation_meta.get("theta", 70)),
                    method=self._form_values["Način polaganja"].get().strip(),
                    loaded_cores=loaded_cores_value,
                    S=s_value if s_value is not None else 1.0,
                    T=t_value if t_value is not None else 1.0,
                    limit_pct=self.DROP_LIMIT_KEYS.get(self._form_values["Ключ ΔU"].get()),
                    current_area=self._try_parse_float(self._form_values["Presek, mm²"].get())
                    or 0,
                    n_parallel=n_parallel,
                    icalc_total=self._last_icalc or 0,
                )
                recommendations = recs or []
            rec_var.set("\n".join(recommendations) if recommendations else "—")

    def _sum_drop_for_circuit(self, circuit: str) -> float:
        if not circuit:
            return 0.0
        total = 0.0
        for row in self._table_data:
            if row.get("Strujni krug", "").strip() == circuit:
                try:
                    total += float(str(row.get("ΔU %", "0")).replace(",", "."))
                except (TypeError, ValueError):
                    continue
        return total

    def _sum_drop_chain_ending_at(self, circuit: str, form_od: str) -> float:
        """Сумма ΔU % только по цепочке участков, продолжающейся до текущего: считаем строки таблицы,
        у которых «До» совпадает с «От» следующего (продолжение кабеля). Обход назад от form_od.
        Если текущее «От» ни с чьим «До» не совпадает — параллельная ветка, возвращаем 0."""
        if not circuit or not form_od:
            return 0.0
        od_var = self._form_values.get("Deonica OD")
        do_var = self._form_values.get("Deonica DO")
        length_var = self._form_values.get("Dužina L, m")
        area_var = self._form_values.get("Presek, mm²")
        form_do = do_var.get().strip() if do_var else ""
        form_l_norm = (length_var.get().strip().replace(",", ".") if length_var else "") or ""
        form_area_norm = (area_var.get().strip().replace(",", ".") if area_var else "") or ""
        total = 0.0
        visited = set()
        current = form_od.strip()
        while current and current not in visited:
            visited.add(current)
            found = None
            for row in self._table_data:
                if row.get("Strujni krug", "").strip() != circuit:
                    continue
                row_do = str(row.get("DO", "")).strip()
                if row_do != current:
                    continue
                row_od = str(row.get("OD", "")).strip()
                row_l = str(row.get("L", "")).replace(",", ".").strip()
                row_area = str(row.get("Presek", "")).replace(",", ".").strip()
                if form_od == row_od and form_do == row_do and form_l_norm == row_l and form_area_norm == row_area:
                    continue
                found = row
                break
            if not found:
                break
            try:
                total += float(str(found.get("ΔU %", "0")).replace(",", "."))
            except (TypeError, ValueError):
                pass
            current = str(found.get("OD", "")).strip()
        return total

    def select_optimal_parameters(self) -> None:
        self._update_intermediate_results()

        insulation_label = self._form_values["Tip-IZOLACIJE"].get()
        insulation_meta = self.INSULATION_META.get(insulation_label)
        if not insulation_meta:
            messagebox.showerror(self._("title.iec"), self._("error.insulation_not_selected"))
            return

        conductor = self._form_values["Tip-PROVODNIKA"].get()
        method = self._form_values["Način polaganja"].get().strip()

        loaded_cores, circuits_count, n_parallel, count_err = self._validate_counts()
        if count_err:
            messagebox.showerror(self._("title.iec"), self._(count_err))
            return

        pi = self._parse_float(self._form_values["Pi, W"].get(), "Pi, W")
        if pi is None:
            return
        kj = self._parse_float(self._form_values["Kj"].get(), "Kj")
        if kj is None:
            return
        eta = self._parse_float(self._form_values["η"].get(), "η")
        if eta is None:
            return
        if not (0 < eta <= 1):
            messagebox.showerror(self._("title.iec"), self._("error.eta_range"))
            return

        cos_phi = self._parse_float(self._form_values["cos φ"].get(), "cos φ")
        if cos_phi is None or not (0 < cos_phi <= 1):
            messagebox.showerror(self._("title.iec"), self._("error.cos_phi_range"))
            return

        length = self._parse_float(self._form_values["Dužina L, m"].get(), "Dužina L, m")
        if length is None or length < 0:
            messagebox.showerror(self._("title.iec"), self._("error.length_non_neg"))
            return

        voltage_str = self._form_values["U"].get()
        try:
            voltage_value = int(voltage_str)
        except (TypeError, ValueError):
            messagebox.showerror(self._("title.iec"), self._("error.voltage_required"))
            return

        medium_key = self._medium_selected_key
        temperature = self._try_parse_float(self._form_values["Температура, °C"].get())
        t_coeff = 1.0
        if temperature is not None:
            temp_factor = self._lookup_temperature_factor(insulation_meta["key"], medium_key, temperature)
            if temp_factor is None or temp_factor <= 0:
                messagebox.showerror(self._("title.iec"), self._("error.temperature_range_kt"))
                return
            t_coeff = temp_factor

        effective_circuits = max(1, circuits_count + n_parallel - 1) if n_parallel > 1 else circuits_count
        s_coeff = self._lookup_group_factor(effective_circuits)

        pj = pi * kj
        phase_factor = 2.0 if loaded_cores == 2 else math.sqrt(3)
        denominator = phase_factor * voltage_value * cos_phi
        if not denominator:
            messagebox.showerror(self._("title.iec"), self._("error.division_zero"))
            return
        icalc_total = (pj / eta) / denominator

        if icalc_total <= 0:
            messagebox.showerror(self._("title.iec"), self._("error.icalc_positive"))
            return

        limit_pct = self.DROP_LIMIT_KEYS.get(self._form_values["Ключ ΔU"].get())
        k_value = self._try_parse_float(self._form_values["k"].get())
        if k_value is None or k_value <= 0:
            k_value = 1.45

        breaker_values: list[tuple[str, float]] = []
        for item in self.STANDARD_BREAKER_RATINGS:
            item = item.strip()
            if not item:
                continue
            try:
                breaker_values.append((item, float(item.replace(",", "."))))
            except ValueError:
                continue

        if not breaker_values:
            messagebox.showerror(self._("title.iec"), self._("error.breaker_ratings_required"))
            return

        insulation_key = insulation_meta["key"]
        insulation_theta = insulation_meta["theta"]

        success_combo: tuple[float, str, float, float] | None = None
        fallback_candidates: list[tuple[float, float, float, float, float, str]] = []

        for area_candidate in self.STANDARD_SECTIONS:
            if area_candidate <= 0:
                continue
            iz_base = self._lookup_ampacity(insulation_key, conductor, method, area_candidate, loaded_cores)
            if not iz_base:
                continue
            iz_one = iz_base * s_coeff * t_coeff
            iz_total = iz_one * max(n_parallel, 1)
            if iz_total <= 0:
                continue
            drop_val = self._drop_pct(
                voltage_value,
                cos_phi,
                length,
                conductor,
                insulation_theta,
                area_candidate,
                method,
                loaded_cores,
                n_parallel,
                icalc_total,
            )
            for breaker_str, breaker_value in breaker_values:
                if breaker_value <= 0:
                    continue
                within_current = icalc_total <= breaker_value <= iz_total
                drop_ok = limit_pct is None or drop_val <= limit_pct
                i2_value = breaker_value * k_value
                protection_ok = i2_value <= 1.45 * iz_total
                if within_current and drop_ok and protection_ok:
                    success_combo = (area_candidate, breaker_str, iz_total, drop_val)
                    break

                over_in_low = max(0.0, icalc_total - breaker_value)
                over_in_high = max(0.0, breaker_value - iz_total)
                over_iz = max(0.0, icalc_total - iz_total)
                over_drop = max(0.0, drop_val - (limit_pct or drop_val)) if limit_pct is not None else 0.0
                over_i2 = max(0.0, i2_value - 1.45 * iz_total)
                metric = over_in_low + over_in_high + over_iz + over_drop + over_i2
                fallback_candidates.append(
                    (
                        metric,
                        area_candidate,
                        breaker_value,
                        drop_val,
                        iz_total,
                        breaker_str,
                    )
                )
            if success_combo:
                break

        if success_combo:
            area_candidate, breaker_str, iz_total, drop_val = success_combo
            if float(area_candidate).is_integer():
                area_text = str(int(area_candidate))
            else:
                area_text = str(area_candidate)
            self._form_values["Presek, mm²"].set(area_text)
            self._form_values["In, A"].set(breaker_str)
            self._update_intermediate_results()
            return

        if not fallback_candidates:
            messagebox.showinfo(self._("title.iec"), self._("message.select_fail") + self._("message.no_combinations"))
            return

        fallback_candidates.sort(key=lambda item: (item[0], item[1], item[2]))
        top_items = fallback_candidates[:3]
        lines: list[str] = []
        for metric, area_candidate, breaker_value, drop_val, iz_total, breaker_str in top_items:
            digits = 0 if float(area_candidate).is_integer() else 1
            area_display = self._fmt(area_candidate, digits=digits)
            iz_display = self._fmt(iz_total, digits=0)
            drop_display = self._fmt(drop_val)
            issues: list[str] = []
            if limit_pct is not None and drop_val > limit_pct:
                issues.append(f"ΔU +{self._fmt(drop_val - limit_pct)}%")
            if breaker_value < icalc_total:
                issues.append(f"In < Ib на {self._fmt(icalc_total - breaker_value)} A")
            if breaker_value > iz_total:
                issues.append(f"In > Iz_tot на {self._fmt(breaker_value - iz_total)} A")
            if breaker_value * k_value > 1.45 * iz_total:
                issues.append(f"I2>{self._fmt(1.45 * iz_total)} A")
            issue_text = "; ".join(issues) if issues else "минимальные отклонения"
            lines.append(
                f"• {area_display} мм² / In={breaker_str} A → Iz_tot≈{iz_display} A, ΔU≈{drop_display}% ({issue_text})"
            )

        message = self._("message.select_fail") + "\n".join(lines)
        messagebox.showinfo(self._("title.iec"), message)

    def remove_selected_row(self) -> None:
        if not hasattr(self, "tree"):
            return
        selected = self.tree.selection()
        if not selected:
            return
        indexed = sorted((self.tree.index(item), item) for item in selected)
        for _, item in indexed:
            self.tree.delete(item)
        for index, _ in reversed(indexed):
            if 0 <= index < len(self._table_data):
                del self._table_data[index]
        self._update_intermediate_results()

    def load_selected_row(self) -> None:
        if not hasattr(self, "tree"):
            return
        selected = self.tree.selection()
        if not selected:
            return
        row_index = self.tree.index(selected[0])
        if not (0 <= row_index < len(self._table_data)):
            return
        row = self._table_data[row_index]

        self._loading_row_into_form = True
        field_map = {
            "Strujni krug": "Strujni krug",
            "Deonica OD": "OD",
            "Deonica DO": "DO",
            "Tip-IZOLACIJE": "E",
            "Tip-PROVODNIKA": "F",
            "Oznaka-tip-KABLA": "G",
            "Pi, W": "Pi",
            "Kj": "Kj",
            "η": "η",
            "U": "U",
            "cos φ": "cosφ",
            "Dužina L, m": "L",
            "Presek, mm²": "Presek",
            "Način polaganja": "Način polaganja",
            "Нагруженные жилы (nž)": "nž",
            "Кабелей в группе (для S)": "Кабелей в группе (S)",
            "Параллельные кабели (n∥)": "n∥",
            "In, A": "In [A]",
            "k": "k",
            "Ключ ΔU": "Ключ",
        }

        current_insulation = self._form_values.get("Tip-IZOLACIJE")
        current_insulation_value = current_insulation.get() if current_insulation else ""

        for field, column in field_map.items():
            if field not in self._form_values:
                continue
            value = str(row.get(column, ""))
            if field == "Presek, mm²" and value:
                numeric = self._try_parse_float(value)
                if numeric is not None:
                    value = str(int(numeric)) if float(numeric).is_integer() else str(numeric)
            if field in {"Pi, W", "Kj", "η", "cos φ", "Dužina L, m"} and value:
                numeric = self._try_parse_float(value)
                if numeric is not None:
                    value = str(numeric)
            widget = self._input_widgets.get(field)
            if isinstance(widget, ttk.Combobox):
                current_values = list(widget.cget("values"))
                if value and value not in current_values:
                    current_values.append(value)
                    widget.configure(values=current_values)
                    self._combobox_values[field] = current_values
            if field == "Tip-IZOLACIJE" and value:
                if value in self.INSULATION_META:
                    self._form_values[field].set(value)
                else:
                    self._form_values[field].set(current_insulation_value)
            else:
                self._form_values[field].set(value)

        self._loading_row_into_form = False
        self._update_intermediate_results()

    def add_row(self) -> None:
        self._update_intermediate_results()

        strujni_krug = self._form_values["Strujni krug"].get().strip()
        od = self._form_values["Deonica OD"].get().strip()
        do = self._form_values["Deonica DO"].get().strip()
        insulation_label = self._form_values["Tip-IZOLACIJE"].get()
        conductor = self._form_values["Tip-PROVODNIKA"].get()
        cable = self._form_values["Oznaka-tip-KABLA"].get().strip()
        laying = self._form_values["Način polaganja"].get().strip()
        voltage = self._form_values["U"].get()
        drop_key = self._form_values["Ключ ΔU"].get()

        pi = self._parse_float(self._form_values["Pi, W"].get(), "Pi, W")
        if pi is None:
            return
        kj = self._parse_float(self._form_values["Kj"].get(), "Kj")
        if kj is None:
            return
        eta = self._parse_float(self._form_values["η"].get(), "η")
        if eta is None:
            return
        if eta <= 0 or eta > 1:
            messagebox.showerror(self._("title.error_input"), self._("error.eta_range_field"))
            return
        cos_phi = self._parse_float(self._form_values["cos φ"].get(), "cos φ")
        if cos_phi is None or cos_phi <= 0 or abs(cos_phi) > 1:
            messagebox.showerror(self._("title.error_input"), self._("error.cos_phi_range_field"))
            return
        length = self._parse_float(self._form_values["Dužina L, m"].get(), "Dužina L, m")
        if length is None:
            return
        area = self._parse_float(self._form_values["Presek, mm²"].get(), "Presek, mm²")
        if area is None or area == 0:
            messagebox.showerror(self._("title.error_input"), self._("error.area_positive"))
            return

        insulation_meta = self.INSULATION_META.get(insulation_label)
        if insulation_meta is None:
            messagebox.showerror(self._("title.error"), self._("error.insulation_unknown"))
            return

        loaded_cores, circuits_count, n_parallel, count_err = self._validate_counts()
        if count_err:
            if count_err == "error.loaded_cores_2_3":
                messagebox.showerror(self._("title.error_input"), self._("error.loaded_cores_2_3_field"))
            elif count_err == "error.circuits_min_1":
                messagebox.showerror(self._("title.error_input"), self._("error.circuits_1_20"))
            else:
                messagebox.showerror(self._("title.error_input"), self._("error.parallel_1_6"))
            return

        if self._last_result is None:
            messagebox.showerror(
                self._("title.error_input"),
                self._("error.add_row_no_result"),
            )
            return

        res = self._last_result
        limit_delta = self.DROP_LIMIT_KEYS.get(drop_key, 0.0)
        existing_drop = self._sum_drop_chain_ending_at(strujni_krug, od)
        total_drop = existing_drop + res["delta_u"]

        if res["ampacity_ok"] == "NE" or res["drop_ok"] == "NE":
            rec_msg = (
                self._intermediate_vars.get("Рекомендации").get()
                if "Рекомендации" in self._intermediate_vars
                else ""
            )
            if rec_msg and rec_msg.strip() != "—":
                messagebox.showinfo(self._("dialog.recommendations.title"), rec_msg)

        row_data = {
            "Strujni krug": strujni_krug,
            "OD": od,
            "DO": do,
            "E": insulation_label,
            "F": conductor,
            "G": cable,
            "nž": str(loaded_cores),
            "n∥": str(n_parallel),
            "Кабелей в группе (S)": str(circuits_count),
            "Pi": self._fmt(pi),
            "Kj": self._fmt(kj),
            "η": self._fmt(eta, digits=3),
            "Pj": self._fmt(res["pj"]),
            "U": voltage,
            "cosφ": self._fmt(cos_phi, digits=3),
            "L": self._fmt(length),
            "Presek": self._fmt(area),
            "Način polaganja": laying,
            "S": self._fmt(res["s_coeff"]),
            "T": self._fmt(res["t_coeff"]),
            "In [A]": self._fmt(res["in_value"]) if res.get("in_value") is not None else "",
            "k": self._fmt(res["k_value"]) if res.get("k_value") is not None else "",
            "I2 [A]": self._fmt(res["i2_value"]) if res.get("i2_value") is not None else "",
            "Icalc [A]": self._fmt(res["icalc_total"], digits=3),
            "R_base [Ω/km]": self._fmt(res["r_per_km"], digits=3),
            "ϭ": self._fmt(res["sigma"], digits=2) if res.get("sigma") is not None else "—",
            "Iz [A]": self._fmt(res["iz_one"]) if res.get("iz_one") is not None else "—",
            "ΔU %": self._fmt(res["delta_u"]),
            "Ukupni ΔU %": self._fmt(total_drop),
            "Limit ΔU %": self._fmt(limit_delta),
            "По току": res["ampacity_ok"],
            "По ΔU": res["drop_ok"],
            "Защита": res["protection_status"],
            "Ключ": drop_key,
            "Совместимость IEC": res["compatibility_status"],
        }

        values = [row_data[column] for column in self.TREE_COLUMNS]
        self.tree.insert("", tk.END, values=values)
        self._table_data.append(row_data)

    def clear_table(self) -> None:
        for item in self.tree.get_children():
            self.tree.delete(item)
        self._table_data.clear()
        self._update_intermediate_results()

    def save_project(self) -> None:
        file_path = filedialog.asksaveasfilename(
            title="Сохранить проект",
            defaultextension=".json",
            filetypes=[("Файл проекта", "*.json"), ("Все файлы", "*.*")],
        )
        if not file_path:
            return

        data = {
            "form": {name: var.get() for name, var in self._form_values.items()},
            "table": self._table_data,
        }

        try:
            with open(file_path, "w", encoding="utf-8") as handle:
                json.dump(data, handle, ensure_ascii=False, indent=2)
        except OSError as exc:
            messagebox.showerror(self._("title.error"), self._m("error.save_project_failed", exc))
            logging.error("Ошибка сохранения проекта '%s': %s", file_path, exc)
        else:
            messagebox.showinfo(self._("title.save"), self._("message.save_ok"))

    def load_project(self) -> None:
        file_path = filedialog.askopenfilename(
            title=self._("dialog.open_project"),
            defaultextension=".json",
            filetypes=[(self._("filetype.project"), "*.json"), (self._("filetype.all"), "*.*")],
        )
        if not file_path:
            return

        try:
            with open(file_path, "r", encoding="utf-8") as handle:
                payload = json.load(handle)
        except (OSError, json.JSONDecodeError) as exc:
            messagebox.showerror(self._("title.error"), self._m("error.load_project_failed", exc))
            logging.error("Ошибка загрузки проекта '%s': %s", file_path, exc)
            return

        form_data = payload.get("form", {})
        table_data = payload.get("table", [])

        self._loading_project = True
        try:
            for name, value in form_data.items():
                if name in self._form_values:
                    if name == "Среда для Т":
                        self._set_medium_from_value(str(value))
                    else:
                        self._form_values[name].set(str(value))
                    widget = self._input_widgets.get(name)
                    if isinstance(widget, ttk.Combobox):
                        current_values = list(widget.cget("values"))
                        display_value = str(value)
                        if display_value not in current_values and display_value != "":
                            current_values.append(display_value)
                            widget.configure(values=current_values)
                            self._combobox_values[name] = current_values

            self.clear_table()

            for row in table_data:
                if not isinstance(row, dict):
                    continue
                normalized = {column: str(row.get(column, "")) for column in self.TREE_COLUMNS}
                self._table_data.append(normalized)
                values = [normalized[column] for column in self.TREE_COLUMNS]
                self.tree.insert("", tk.END, values=values)

            self._loading_project = False
            self._update_intermediate_results()
        finally:
            self._loading_project = False

    def export_to_excel(self) -> None:
        if not self._table_data:
            messagebox.showinfo(self._("title.export"), self._("message.export_no_data"))
            return

        file_path = filedialog.asksaveasfilename(
            title=self._("dialog.save_as"),
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
        )
        if not file_path:
            return

        try:
            self._write_workbook(file_path)
            messagebox.showinfo(self._("title.export"), self._("message.export_ok"))
        except (OSError, ValueError) as exc:
            messagebox.showerror(self._("title.error"), self._m("error.export_failed", exc))
            logging.error("Ошибка экспорта Excel '%s': %s", file_path, exc)

    def _write_workbook(self, file_path: str) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Proračuni"

        num_cols = len(self.TREE_COLUMNS)
        lang = self._language.get()
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        row_num = 1
        title_cell = worksheet.cell(row=row_num, column=1, value=self._("export.doc_title"))
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        if num_cols > 1:
            worksheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=num_cols)
        row_num += 1

        subtitle_cell = worksheet.cell(row=row_num, column=1, value=self._("export.subtitle"))
        subtitle_cell.font = Font(bold=True, size=11)
        if num_cols > 1:
            worksheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=num_cols)
        row_num += 1

        date_cell = worksheet.cell(row=row_num, column=1, value=datetime.now().strftime("%d.%m.%Y"))
        date_cell.font = Font(italic=True)
        row_num += 1

        std_cell = worksheet.cell(row=row_num, column=1, value=self._("export.standard"))
        std_cell.alignment = Alignment(wrap_text=True)
        std_cell.font = Font(size=9)
        if num_cols > 1:
            worksheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=num_cols)
        row_num += 2

        table_header_row = row_num
        header_row = [self._(self.TREE_COLUMN_KEYS.get(col, col)) for col in self.TREE_COLUMNS]
        for col_idx, value in enumerate(header_row, start=1):
            cell = worksheet.cell(row=table_header_row, column=col_idx, value=value)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row_num += 1

        for row in self._table_data:
            for col_idx, column in enumerate(self.TREE_COLUMNS, start=1):
                raw = row.get(column, "")
                if isinstance(raw, (int, float)):
                    cell_val: typing.Any = raw
                else:
                    text = str(raw)
                    number = self._try_parse_float(text)
                    cell_val = number if number is not None and text.strip() not in {"", "—"} else text
                cell = worksheet.cell(row=row_num, column=col_idx, value=cell_val)
                cell.border = thin_border
            row_num += 1

        data_end_row = row_num - 1
        worksheet.auto_filter.ref = f"A{table_header_row}:{get_column_letter(num_cols)}{data_end_row}"
        worksheet.freeze_panes = f"A{table_header_row + 1}"

        row_num += 1
        legend_title_cell = worksheet.cell(row=row_num, column=1, value=self._("export.legend_title"))
        legend_title_cell.font = Font(bold=True, size=11)
        row_num += 1

        for col in self.TREE_COLUMNS:
            key = self.TREE_COLUMN_KEYS.get(col, col)
            label = self._(key)
            desc_key = self.COLUMN_DESC_KEYS.get(key, key)
            if desc_key.startswith("label."):
                tt = self.TOOLTIPS.get(desc_key, {})
                desc = tt.get(lang) or tt.get(self.DEFAULT_LANGUAGE) or (next(iter(tt.values())) if tt else "")
            else:
                desc = self._(desc_key)
            if desc:
                legend_cell = worksheet.cell(row=row_num, column=1, value=f"{label} — {desc}")
                legend_cell.alignment = Alignment(wrap_text=False)
                row_num += 1

        for col_idx in range(1, num_cols + 1):
            max_length = 12
            for r in range(table_header_row, data_end_row + 1):
                cell = worksheet.cell(row=r, column=col_idx)
                if cell.value is not None:
                    max_length = max(max_length, min(len(str(cell.value)) + 2, 50))
            worksheet.column_dimensions[get_column_letter(col_idx)].width = max_length

        workbook.save(file_path)


def main() -> None:
    app = CableCalcApp()
    app.mainloop()


if __name__ == "__main__":
    main()
